"""
Recipe Checker - Core Logic
Analyzes recipes for food allergies and dietary restrictions.
"""

import re
import openpyxl
from collections import defaultdict


def load_food_map(excel_path='Food_Map_Levels.xlsx'):
    """Load food → risk level mapping from Excel file."""
    food_map = {}
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] is not None:
                food_item = str(row[0]).strip().lower()
                level = int(row[1])
                notes = str(row[2]) if row[2] else ''
                food_map[food_item] = {'level': level, 'notes': notes}
    
    except FileNotFoundError:
        print(f"Error: Could not find {excel_path}")
        return {}
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return {}
    
    return food_map


def clean_food_item_name(food_item):
    """Clean food item name by removing measurements and extra whitespace."""
    food_item_clean = str(food_item).strip()
    # Remove checkbox symbols and other special characters
    food_item_clean = re.sub(r'^[▢▣☐☑☒✓✗]\s*', '', food_item_clean)
    # Remove leading numbers, fractions, and measurement words
    # Pattern matches: numbers (including fractions like 1/2, 1/4), optional hyphens, and measurement words
    food_item_clean = re.sub(
        r'^(?:\d+(?:\/\d+)?(?:-\d+(?:\/\d+)?)?\s*)?(?:cups?|tbsp\.?|tablespoons?|tsp\.?|teaspoons?|lb\.?|lbs\.?|pound|pounds|oz\.?|ounce|ounces|gram|grams|g\.?|kg\.?|kilogram|kilograms|ml\.?|milliliter|milliliters|l\.?|liter|liters|clove|cloves|piece|pieces|slice|slices|can|cans|bunch|bunches|pinch|pinches|dash|dashes|large|small|medium|extra-virgin|virgin)\s+',
        '',
        food_item_clean,
        flags=re.IGNORECASE
    )
    food_item_clean = re.sub(r'\s+(?:to\s+)?taste.*$', '', food_item_clean, flags=re.IGNORECASE)
    food_item_clean = re.sub(r'\s+as\s+needed.*$', '', food_item_clean, flags=re.IGNORECASE)
    food_item_clean = re.sub(r'\s+\(.*?\)\s*$', '', food_item_clean)
    food_item_clean = ' '.join(food_item_clean.split()).strip()
    food_item_clean = re.sub(r'[:;.,!?]+$', '', food_item_clean).strip()
    
    return food_item_clean


def add_food_item(excel_path, food_item, level, notes=''):
    """Add a new food item to the Excel file. Measurements are automatically removed."""
    try:
        food_item_clean = clean_food_item_name(food_item)
        
        if not food_item_clean:
            print(f"Error: Food item name is empty after cleaning")
            return False
        
        food_item_lower = food_item_clean.lower()
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value:
                existing_item = clean_food_item_name(str(row[0].value))
                if existing_item.lower() == food_item_lower:
                    row[0].value = food_item_clean
                    row[1].value = int(level)
                    row[2].value = str(notes) if notes else ''
                    wb.save(excel_path)
                    return True
        
        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1, value=food_item_clean)
        ws.cell(row=next_row, column=2, value=int(level))
        ws.cell(row=next_row, column=3, value=str(notes) if notes else '')
        
        wb.save(excel_path)
        return True
        
    except Exception as e:
        print(f"Error adding food item: {e}")
        return False


def extract_all_ingredients(recipe_text):
    """Extract all ingredients from recipe text maintaining 1-to-1 mapping."""
    lines = recipe_text.split('\n')
    non_empty_lines = [line.strip() for line in lines if line.strip()]
    use_line_breaks = len(non_empty_lines) > 1
    
    all_ingredients = []
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        if line.isupper() and len(line) < 50:
            continue
        
        if ':' in line:
            parts = line.split(':', 1)
            if len(parts) == 2:
                header = parts[0].strip().lower()
                skip_headers = {'ingredient', 'ingredients', 'recipe', 'instructions', 'directions', 'method', 'steps'}
                if header in skip_headers or any(header.startswith(word) for word in skip_headers):
                    line = parts[1].strip()
                    if not line:
                        continue
        
        if use_line_breaks:
            all_items = [line]
        else:
            all_items = [item.strip() for item in line.split(',')]
        
        processed_items = []
        for item in all_items:
            item = item.strip()
            if not item:
                continue
            
            item = re.sub(r'^(or|Or)\s+', '', item, flags=re.IGNORECASE).strip()
            if not item:
                continue
            
            paren_pattern = r'\([^)]*\)'
            paren_matches = re.findall(paren_pattern, item)
            
            if paren_matches:
                placeholders = []
                temp_item = item
                for i, match in enumerate(paren_matches):
                    placeholder = f"__PAREN_{i}__"
                    placeholders.append(match)
                    temp_item = temp_item.replace(match, placeholder, 1)
                
                or_separated = re.split(r'\s+or\s+', temp_item, flags=re.IGNORECASE)
                
                for subitem in or_separated:
                    restored = subitem.strip()
                    for i, placeholder in enumerate([f"__PAREN_{j}__" for j in range(len(placeholders))]):
                        if placeholder in restored:
                            restored = restored.replace(placeholder, placeholders[i])
                    if restored:
                        processed_items.append(restored)
            else:
                or_separated = re.split(r'\s+or\s+', item, flags=re.IGNORECASE)
                processed_items.extend([subitem.strip() for subitem in or_separated if subitem.strip()])
        
        for ingredient_raw in processed_items:
            ingredient_raw = ingredient_raw.strip()
            if not ingredient_raw:
                continue
            
            ingredient_clean = re.sub(
                r'^\d+\s*(?:cups?|tbsp|tablespoons?|tsp|teaspoons?|lb|lbs|pound|pounds|oz|ounce|ounces|gram|grams|kg|kilogram|kilograms|ml|milliliter|milliliters|l|liter|liters|clove|cloves|piece|pieces|slice|slices|can|cans|bunch|bunches|pinch|pinches|dash|dashes)\s+',
                '',
                ingredient_raw,
                flags=re.IGNORECASE
            )
            ingredient_clean = re.sub(r'\s+(?:to\s+)?taste.*$', '', ingredient_clean, flags=re.IGNORECASE)
            ingredient_clean = re.sub(r'\s+as\s+needed.*$', '', ingredient_clean, flags=re.IGNORECASE)
            ingredient_clean = re.sub(r'\s+\(.*?\)\s*$', '', ingredient_clean)
            ingredient_clean = ' '.join(ingredient_clean.split()).strip()
            ingredient_clean = re.sub(r'[:;.,!?]+$', '', ingredient_clean).strip()
            
            if ingredient_clean and len(ingredient_clean) > 1:
                ingredient_lower = ingredient_clean.lower()
                skip_words = {'ingredient', 'ingredients', 'recipe', 'instructions', 'directions', 'method', 'steps', 'save', 'print','bookmark','bookmarks','printprint','see all nutritional information'}
                
                if ingredient_lower not in skip_words and not any(ingredient_lower.startswith(word + ' ') for word in skip_words):
                    all_ingredients.append(ingredient_lower)
    
    return all_ingredients


def parse_recipe(recipe_text, food_map):
    """Parse recipe text and match ingredients against food map."""
    recipe_lower = recipe_text.lower()
    found_items = defaultdict(lambda: {'level': 0, 'notes': '', 'count': 0, 'matched': True})
    all_ingredients_dict = defaultdict(lambda: {'matched': False, 'level': None, 'notes': '', 'count': 0})
    
    all_ingredients = extract_all_ingredients(recipe_text)
    
    for ingredient in all_ingredients:
        all_ingredients_dict[ingredient] = {
            'matched': False,
            'level': None,
            'notes': '',
            'count': 1
        }
    
    def extract_core_ingredient(text):
        """Extract core ingredient name by removing measurements."""
        text_lower = text.lower()
        text_lower = re.sub(
            r'^\d+\s*(?:cups?|tbsp|tablespoons?|tsp|teaspoons?|lb|lbs|pound|pounds|oz|ounce|ounces|gram|grams|kg|kilogram|kilograms|ml|milliliter|milliliters|l|liter|liters|clove|cloves|piece|pieces|slice|slices|can|cans|bunch|bunches|pinch|pinches|dash|dashes)\s+',
            '',
            text_lower,
            flags=re.IGNORECASE
        )
        return text_lower.strip()
    
    for ingredient in all_ingredients:
        ingredient_lower = ingredient.lower()
        ingredient_core = extract_core_ingredient(ingredient)
        
        if 'corn' in ingredient_lower:
            all_ingredients_dict[ingredient] = {
                'matched': True,
                'level': 3,
                'notes': 'Critical - contains corn',
                'count': 1
            }
            found_items[ingredient] = {
                'level': 3,
                'notes': 'Critical - contains corn',
                'count': 1,
                'matched': True
            }
            continue
        
        if ingredient_lower in food_map:
            info = food_map[ingredient_lower]
            all_ingredients_dict[ingredient] = {
                'matched': True,
                'level': info['level'],
                'notes': info['notes'],
                'count': 1
            }
            found_items[ingredient_lower] = {
                'level': info['level'],
                'notes': info['notes'],
                'count': 1,
                'matched': True
            }
            continue
        
        matched = False
        for food_item, info in food_map.items():
            food_item_lower = food_item.lower()
            food_item_core = extract_core_ingredient(food_item)
            
            if 'corn' in food_item_lower:
                continue
            
            pattern = r'\b' + re.escape(food_item_lower) + r'\b'
            if re.search(pattern, ingredient_lower, re.IGNORECASE):
                all_ingredients_dict[ingredient] = {
                    'matched': True,
                    'level': info['level'],
                    'notes': info['notes'],
                    'count': 1
                }
                if food_item_lower not in found_items:
                    found_items[food_item_lower] = {
                        'level': info['level'],
                        'notes': info['notes'],
                        'count': 1,
                        'matched': True
                    }
                matched = True
                break
            
            if ingredient_core and food_item_core:
                if ingredient_core == food_item_core:
                    all_ingredients_dict[ingredient] = {
                        'matched': True,
                        'level': info['level'],
                        'notes': info['notes'],
                        'count': 1
                    }
                    if food_item_lower not in found_items:
                        found_items[food_item_lower] = {
                            'level': info['level'],
                            'notes': info['notes'],
                            'count': 1,
                            'matched': True
                        }
                    matched = True
                    break
                elif ingredient_core in food_item_core or food_item_core in ingredient_core:
                    shorter = ingredient_core if len(ingredient_core) < len(food_item_core) else food_item_core
                    longer = food_item_core if len(ingredient_core) < len(food_item_core) else ingredient_core
                    if re.search(r'\b' + re.escape(shorter) + r'\b', longer, re.IGNORECASE):
                        all_ingredients_dict[ingredient] = {
                            'matched': True,
                            'level': info['level'],
                            'notes': info['notes'],
                            'count': 1
                        }
                        if food_item_lower not in found_items:
                            found_items[food_item_lower] = {
                                'level': info['level'],
                                'notes': info['notes'],
                                'count': 1,
                                'matched': True
                            }
                        matched = True
                        break
    
    return dict(found_items), dict(all_ingredients_dict)


def categorize_foods(found_items):
    """Categorize found foods by risk level."""
    categorized = {0: [], 1: [], 2: [], 3: []}
    
    for food_item, info in found_items.items():
        level = info['level']
        categorized[level].append((food_item, info))
    
    return categorized


def calculate_total_risk_score(found_items):
    """Calculate total risk score by summing all risk levels."""
    total_score = 0
    for food_item, info in found_items.items():
        total_score += info['level'] * info['count']
    return total_score


def analyze_recipe(recipe_text, excel_path='Food_Map_Levels.xlsx'):
    """Main analysis function."""
    food_map = load_food_map(excel_path)
    
    if not food_map:
        return {
            'error': 'Could not load food map',
            'categorized': {},
            'total_score': 0,
            'all_ingredients': {}
        }
    
    found_items, all_ingredients = parse_recipe(recipe_text, food_map)
    categorized = categorize_foods(found_items)
    total_score = calculate_total_risk_score(found_items)
    
    return {
        'found_items': found_items,
        'categorized': categorized,
        'total_score': total_score,
        'food_map': food_map,
        'all_ingredients': all_ingredients
    }

